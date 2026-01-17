import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configuração de estilos constante (facilita manutenção)
COLORS_ALA = {
    '4ª': {'bg': 'FF0000', 'font': 'FFFFFF'},  # Vermelho / Branco
    '3ª': {'bg': '00FF00', 'font': 'FFFFFF'},  # Verde / Branco
    '2ª': {'bg': 'FFFF00', 'font': '000000'},  # Amarelo / Preto
    '1ª': {'bg': '0000FF', 'font': 'FFFFFF'},  # Azul / Branco
}

def aplicar_estilos_celula(ws):
    """Aplica alinhamento, bordas e ajuste automático de colunas."""
    border_style = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in ws.iter_cols(1, ws.max_column):
        nome_coluna = col[0].value # Pega o nome do cabeçalho
        
        # Ajuste de largura padrão
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        
        # --- REGRA ESPECIAL PARA O HISTÓRICO ---
        if nome_coluna == 'Histórico':
            ws.column_dimensions[col[0].column_letter].width = 60 # Largura fixa maior
        else:
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        for cell in col:
            # Se for histórico, permite quebra de linha para o texto não 'vazar'
            if nome_coluna == 'Histórico' and cell.row > 1:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if cell.row > 1: # Bordas apenas nos dados
                cell.border = border_style
            else: # Cabeçalho em negrito
                cell.font = Font(bold=True)

def aplicar_cores_ala(ws):
    """Aplica a formatação condicional baseada na coluna ALA."""
    col_idx = None
    for cell in ws[1]:
        if cell.value == 'ALA':
            col_idx = cell.column_letter
            break
    
    if not col_idx:
        return

    for cell in ws[col_idx][1:]: # Pula o cabeçalho
        style = COLORS_ALA.get(str(cell.value))
        if style:
            cell.fill = PatternFill(start_color=style['bg'], end_color=style['bg'], fill_type="solid")
            cell.font = Font(color=style['font'], bold=(style['font'] == '000000'))

def organizar_planilha(file_path: Path):
    if not file_path.exists():
        print(f"⚠️ Arquivo não encontrado: {file_path}")
        return False

    try:

        file_path.rename(file_path)
    except OSError:
        print(f"❌ Erro: O arquivo {file_path.name} está aberto no Excel. Feche-o e tente novamente.")
        return False

    try:
        # Etapa 1: Pandas (Lógica de dados)
        df = pd.read_excel(file_path)
        col_data = 'Data/hora de criação'
        
        if col_data in df.columns:
            df[col_data] = pd.to_datetime(df[col_data], dayfirst=True, errors='coerce')
            df = df.sort_values(by=col_data, ascending=False) # Mais recentes no topo
            df[col_data] = df[col_data].dt.strftime('%d/%m/%Y %H:%M')
            
            df.to_excel(file_path, index=False)

        # Etapa 2: Openpyxl (Estética)
        wb = load_workbook(file_path)
        ws = wb.active
        
        aplicar_estilos_celula(ws)
        aplicar_cores_ala(ws)
        
        wb.save(file_path)
        print(f"✅ Planilha {file_path.name} organizada e formatada!")
        return True

    except Exception as e:
        print(f"❌ Erro ao organizar {file_path.name}: {e}")
        return False

# Mantém a compatibilidade para rodar o script sozinho
if __name__ == "__main__":
    # Localiza o ROOT_DIR de forma dinâmica para o script solto
    BASE_DIR = Path(__file__).resolve().parent.parent
    target = BASE_DIR / "data" / "chamadas_csv" / "nova_planilha_ocorrencias.xlsx"
    organizar_planilha(target)