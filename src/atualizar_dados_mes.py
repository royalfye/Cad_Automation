import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

ROOT_DIR = Path(__file__).resolve().parent 
if ROOT_DIR.name != "Cad_Automation":
    ROOT_DIR = ROOT_DIR.parent

ASSETS_DIR = ROOT_DIR / "assets"
DATA_DIR = ROOT_DIR / "data" / "chamadas_csv"
def determinar_ala(data_hora_str):
    try:
        data_hora = pd.to_datetime(data_hora_str, format='%d/%m/%Y %H:%M', errors='raise')
        data_referencia = pd.to_datetime('01/01/2025 07:45:00')
        if data_hora.time() < pd.to_datetime('07:45:00').time():
            dias_offset = (data_hora.date() - data_referencia.date()).days - 1
        else:
            dias_offset = (data_hora.date() - data_referencia.date()).days
        sequencia_alas = ['4ª', '1ª', '2ª', '3ª']
        indice_ala = dias_offset % 4
        return sequencia_alas[indice_ala]
    except Exception as e:
        print(f"Erro ao determinar a ala para '{data_hora_str}': {e}")
        return "Indefinida"

def limpar_coluna_natureza(df):
    df['Natureza'] = df['Natureza'].str.replace(r'\s*\(.*$', '', regex=True)
    return df

def formatar_excel(excel_file_path):
    wb = load_workbook(excel_file_path)
    ws = wb.active
    for col in ws.iter_cols(1, ws.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[col[0].column_letter].width = 45
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(excel_file_path)

def atualizar_dados_occurrences(csv_file_path, excel_file_path):

    if csv_file_path.exists():
        try:

            df_novo = pd.read_csv(csv_file_path, encoding='latin1', sep=';', dtype=str)
            df_novo.columns = df_novo.columns.str.strip()

            colunas_necessarias = [
                'Data/hora de criação', 'Data/hora da situação atual', 'Nº chamada', 'Natureza',
                'Local do fato', 'Unidade Responsável', 'Recursos empenhados'
            ]

            for col in colunas_necessarias:
                if col not in df_novo.columns:
                    print(f"Coluna '{col}' não encontrada no CSV.")
                    return

            df_novo['Data/hora de criação'] = pd.to_datetime(
                df_novo['Data/hora de criação'].str.strip(), format='%d/%m/%Y %H:%M', errors='coerce'
            )
            df_novo['Data/hora da situação atual'] = pd.to_datetime(
                df_novo['Data/hora da situação atual'].str.strip(), format='%d/%m/%Y %H:%M', errors='coerce'
            )
            
            df_novo = df_novo.dropna(subset=['Data/hora de criação', 'Data/hora da situação atual'])
            df_novo['Data/hora de criação'] = df_novo['Data/hora de criação'].dt.strftime('%d/%m/%Y %H:%M')
            df_novo['Data/hora da situação atual'] = df_novo['Data/hora da situação atual'].dt.strftime('%d/%m/%Y %H:%M')
            
            df_novo_filtrado = df_novo[colunas_necessarias].copy()
            df_novo_filtrado = limpar_coluna_natureza(df_novo_filtrado)
            df_novo_filtrado['Classe'] = df_novo_filtrado['Natureza'].str.strip().str[0]
            df_novo_filtrado['ALA'] = df_novo_filtrado['Data/hora de criação'].apply(determinar_ala)

            df_novo_filtrado = df_novo_filtrado[[
                'Data/hora de criação', 'Data/hora da situação atual', 'ALA', 'Classe', 'Nº chamada', 
                'Natureza', 'Local do fato', 'Recursos empenhados'
            ]]
            df_novo_filtrado.rename(columns={'Data/hora da situação atual': 'Data/hora final'}, inplace=True)

            if excel_file_path.exists():
                df_existente = pd.read_excel(excel_file_path, sheet_name='Sheet1', dtype=str)
                df_junto = pd.concat([df_existente, df_novo_filtrado]).drop_duplicates(subset=['Nº chamada'], keep='last')
            else:
                df_junto = df_novo_filtrado

            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='w') as writer:
                df_junto.to_excel(writer, sheet_name='Sheet1', index=False)

            formatar_excel(excel_file_path)
            print(f"Total de registros processados: {len(df_junto)}")
        except Exception as e:
            print(f"Erro ao processar arquivo CSV '{csv_file_path}': {e}")
    else:
        print(f"Arquivo CSV não encontrado: {csv_file_path}")

def main():
    # Mapeamento utilizando DATA_DIR e a sintaxe / da pathlib
    arquivos_csv = {
        "Janeiro": DATA_DIR / "ocorrencias_jan_2025.csv",
        "Fevereiro": DATA_DIR / "ocorrencias_fev_2025.csv",
        "Março": DATA_DIR / "ocorrencias_mar_2025.csv",
        "Abril": DATA_DIR / "ocorrencias_abr_2025.csv"
    }
    
    excel_file_path = DATA_DIR / "nova_planilha_ocorrencias.xlsx"

    for mes, caminho_csv in arquivos_csv.items():
        print(f"Processando dados de {mes}...")
        atualizar_dados_occurrences(caminho_csv, excel_file_path)

if __name__ == "__main__":
    main()