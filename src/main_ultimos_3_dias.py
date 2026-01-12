import time
import pyautogui
import pandas as pd
import os
import pygetwindow as gw
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import logging
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parent.parent

ASSETS_DIR = ROOT_DIR / "assets"
DATA_DIR = ROOT_DIR / "data" / "chamadas_csv"

csv_path = DATA_DIR / "ocorrencias_classificadas.csv"
excel_path = DATA_DIR / "nova_planilha_ocorrencias.xlsx"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

def determinar_ala(data_hora_str):
    try:
        data_hora = pd.to_datetime(data_hora_str, format='%d/%m/%Y %H:%M', errors='raise')
        data_referencia = pd.to_datetime('01/01/2025 07:45:00')
        if data_hora.time() < pd.to_datetime('07:45:00').time():
            dias_offset = (data_hora.date() - data_referencia.date()).days - 1
        else:
            dias_offset = (data_hora.date() - data_referencia.date()).days
        sequencia_alas = ['4ª', '1ª', '2ª', '3ª']
        indice_ala = dias_offset % 4
        return sequencia_alas[indice_ala]
    except Exception as e:
        logging.error(f"Erro ao determinar a ala para '{data_hora_str}': {e}")
        return "Indefinida"

def activate_window(window_title):
    try:
        while True:
            try:
                window = gw.getWindowsWithTitle(window_title)[0]
                window.activate()
                window.maximize()
                logging.info(f"Janela ativada: {window_title}")
                break
            except IndexError:
                logging.warning(f"Esperando pela janela: {window_title}")
                time.sleep(1)
    except Exception as e:
        logging.error(f"Erro ao ativar janela: {e}")

def limpar_coluna_natureza(df):
    df['Natureza'] = df['Natureza'].str.replace(r'\s*\(.*$', '', regex=True)
    return df

def formatar_excel(excel_file_path):
    wb = load_workbook(excel_file_path)
    ws = wb.active
    for col in ws.iter_cols(1, ws.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[col[0].column_letter].width = 45
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(excel_file_path)
    logging.info("Excel formatado com sucesso.")

def atualizar_dados_occurrences(csv_file_path, excel_file_path):
    if csv_file_path.exists():
        try:
            df_novo = pd.read_csv(csv_file_path, encoding='latin1', sep=';', dtype=str)
            df_novo.columns = df_novo.columns.str.strip()
            logging.info(f"Colunas lidas do CSV: {df_novo.columns.tolist()}")

            colunas_necessarias = [
                'Data/hora de criação', 'Data/hora da situação atual', 'Nº chamada', 
                'Natureza', 'Local do fato', 'Unidade Responsável', 'Recursos empenhados'
            ]
            for col in colunas_necessarias:
                if col not in df_novo.columns:
                    logging.warning(f"Coluna '{col}' não encontrada no CSV.")
                    return

            df_novo = df_novo[df_novo['Unidade Responsável'].str.contains(r'\(PASSOS\)', na=False)]
            df_novo_filtrado = df_novo[colunas_necessarias].copy()

            df_novo_filtrado['Data/hora de criação'] = pd.to_datetime(
                df_novo_filtrado['Data/hora de criação'], format='%d/%m/%Y %H:%M', errors='coerce'
            ).dt.strftime('%d/%m/%Y %H:%M')

            df_novo_filtrado['Data/hora da situação atual'] = pd.to_datetime(
                df_novo_filtrado['Data/hora da situação atual'], format='%d/%m/%Y %H:%M', errors='coerce'
            ).dt.strftime('%d/%m/%Y %H:%M')

            df_novo_filtrado = limpar_coluna_natureza(df_novo_filtrado)
            df_novo_filtrado['Classe'] = df_novo_filtrado['Natureza'].str.strip().str[0]
            df_novo_filtrado['ALA'] = df_novo_filtrado['Data/hora de criação'].apply(determinar_ala)

            df_novo_filtrado = df_novo_filtrado[['Data/hora de criação', 'Data/hora da situação atual', 'ALA', 
                                                  'Classe', 'Nº chamada', 'Natureza', 'Local do fato', 
                                                  'Recursos empenhados']]
            df_novo_filtrado.rename(columns={'Data/hora da situação atual': 'Data/hora final'}, inplace=True)

            if excel_file_path.exists():
                df_existente = pd.read_excel(excel_file_path, dtype=str)
                df_junto = pd.concat([df_existente, df_novo_filtrado]).drop_duplicates(subset=['Nº chamada'], keep='last')
                df_junto.reset_index(drop=True, inplace=True)
                df_junto.to_excel(excel_file_path, index=False)
            else:
                df_novo_filtrado.to_excel(excel_file_path, index=False)

            formatar_excel(excel_file_path)
            logging.info(f"Total de registros processados: {len(df_novo_filtrado)}")
            for ala in ['1ª', '2ª', '3ª', '4ª']:
                logging.info(f"Registros na {ala} ALA: {(df_novo_filtrado['ALA'] == ala).sum()}")
        except Exception as e:
            logging.error(f"Erro ao processar o arquivo CSV: {e}")
    else:
        logging.error(f"Arquivo CSV não encontrado em: {csv_file_path}")

def main_classificadas():
    window_title = "CAD - Solução de Controle do Atendimento e Despacho de Emergência Policial e de Bombeiros"

    activate_window(window_title)
    time.sleep(1)
    
    logging.info("Iniciando cliques de automação...")
    try:
        pyautogui.click(str(ASSETS_DIR / 'chamadas_button.png'))
        time.sleep(1)
        pyautogui.click(str(ASSETS_DIR / 'pesquisa_button.png'))
        time.sleep(1)

        pyautogui.click(x=402, y=211)
        time.sleep(1)
        
        ultimas24 = pyautogui.locateCenterOnScreen(str(ASSETS_DIR / 'ultimas_24.png'), confidence=0.7)
        if ultimas24:
            pyautogui.click(ultimas24)
        else:
            logging.warning("Imagem 'ultimas_24.png' não encontrada.")
        time.sleep(1)
        
        ultimos3 = pyautogui.locateCenterOnScreen(str(ASSETS_DIR / 'ultimos_3.png'), confidence=0.7)
        if ultimos3:
            pyautogui.click(ultimos3)
        else:
            logging.warning("Imagem 'ultimos_3.png' não encontrada.")
        time.sleep(1)
        
        if pyautogui.locateOnScreen(str(ASSETS_DIR / 'seta_button.png'), confidence=0.7):
            pyautogui.click(str(ASSETS_DIR / 'seta_button.png'))
            time.sleep(1)
        else:
            logging.warning("Botão da seta não encontrado.")
            
        pyautogui.write("PASSOS")
        time.sleep(1)
        
        passos_exibido = pyautogui.locateCenterOnScreen(str(ASSETS_DIR / 'passos_exibido.png'), confidence=0.7)
        if passos_exibido:
            pyautogui.click(passos_exibido)
            time.sleep(1)
        else:
            logging.warning("Palavra 'PASSOS' não encontrada na tela.")

        pyautogui.click(x=1153, y=414)
        time.sleep(3)
        pyautogui.press('f12')
        time.sleep(1)
        # pyautogui.write exige string
        pyautogui.write(str(csv_path))
        time.sleep(1)
        pyautogui.press('enter')    
        pyautogui.hotkey('alt', 'f4')

        try:
            pesquisa_window = gw.getWindowsWithTitle("Pesquisa Chamadas")[0]
            pesquisa_window.activate()
            time.sleep(1)
            pyautogui.hotkey('alt', 'f4')
        except IndexError:
            logging.warning("Janela 'Pesquisa Chamadas' não encontrada.")

        timeout = 30
        start_time = time.time()
        while not csv_path.exists():
            if time.time() - start_time > timeout:
                logging.critical("CSV não foi gerado no tempo esperado!")
                return
            time.sleep(1)

        try:
            ocorrencias_csv_window = gw.getWindowsWithTitle("ocorrencias_classificadas.csv - Excel")[0]
            ocorrencias_csv_window.activate()
            time.sleep(1)
            pyautogui.hotkey('alt', 'f4')
        except IndexError:
            logging.warning("Janela 'ocorrencias_classificadas.csv - Excel' não encontrada.")

        atualizar_dados_occurrences(csv_path, excel_path)
        logging.info("Dados exportados e a nova planilha atualizada e formatada com sucesso!")
        
    except Exception as e:
        logging.error(f"Erro na automação: {e}")

if __name__ == "__main__":
    main_classificadas()